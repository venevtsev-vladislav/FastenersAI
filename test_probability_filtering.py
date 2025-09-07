#!/usr/bin/env python3
"""
Тест фильтрации результатов по вероятности
"""

import asyncio
import sys
import os
import json

# Добавляем родительскую папку в путь для импорта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from pipeline.message_processor import MessagePipeline

async def test_probability_filtering():
    """Тестируем фильтрацию по вероятности"""
    
    print("🔍 ТЕСТ ФИЛЬТРАЦИИ ПО ВЕРОЯТНОСТИ")
    print("=" * 60)
    
    # Тестовые данные с разными вероятностями
    test_search_results = [
        {
            "sku": "10-0010040",
            "name": "Анкер забиваемый латунный М10",
            "smart_probability": 85,  # Высокая вероятность - должна пройти
            "diameter": "M10",
            "material": "латунь"
        },
        {
            "sku": "6-0049020", 
            "name": "Болт DIN 933 кл.пр.8.8 М10х30, цинк",
            "smart_probability": 70,  # Средняя вероятность - должна пройти
            "diameter": "M10",
            "coating": "цинк"
        },
        {
            "sku": "12-0013800",
            "name": "Винт DIN 7985 М4х40, цинк",
            "smart_probability": 45,  # Низкая вероятность - должна быть исключена
            "diameter": "M4",
            "coating": "цинк"
        },
        {
            "sku": "8-0021500",
            "name": "Гайка М8, цинк",
            "smart_probability": 30,  # Очень низкая вероятность - должна быть исключена
            "diameter": "M8"
        },
        {
            "sku": "15-0032000",
            "name": "Шайба М10, цинк",
            "smart_probability": 65,  # Граничная вероятность - должна пройти
            "diameter": "M10"
        }
    ]
    
    try:
        # Создаем pipeline
        pipeline = MessagePipeline(bot=None)
        
        print(f"📊 Исходные данные: {len(test_search_results)} результатов")
        for i, result in enumerate(test_search_results, 1):
            print(f"  {i}. {result['sku']} - вероятность {result['smart_probability']}%")
        
        # Тестируем фильтрацию
        print(f"\n🔍 ТЕСТИРУЕМ ФИЛЬТРАЦИЮ (вероятность >= 60%):")
        filtered_results = pipeline._filter_results_by_probability(test_search_results)
        
        print(f"✅ Результат фильтрации: {len(filtered_results)} результатов")
        for i, result in enumerate(filtered_results, 1):
            print(f"  {i}. {result['sku']} - вероятность {result['smart_probability']}%")
        
        # Проверяем, что исключены правильные результаты
        excluded_skus = [r['sku'] for r in test_search_results if r not in filtered_results]
        included_skus = [r['sku'] for r in filtered_results]
        
        print(f"\n📋 АНАЛИЗ РЕЗУЛЬТАТОВ:")
        print(f"  - Включены: {included_skus}")
        print(f"  - Исключены: {excluded_skus}")
        
        # Проверяем корректность фильтрации
        expected_included = ["10-0010040", "6-0049020", "15-0032000"]  # >= 60%
        expected_excluded = ["12-0013800", "8-0021500"]  # < 60%
        
        success = True
        for sku in expected_included:
            if sku not in included_skus:
                print(f"❌ ОШИБКА: {sku} должен быть включен, но исключен")
                success = False
        
        for sku in expected_excluded:
            if sku not in excluded_skus:
                print(f"❌ ОШИБКА: {sku} должен быть исключен, но включен")
                success = False
        
        if success:
            print(f"✅ ФИЛЬТРАЦИЯ РАБОТАЕТ ПРАВИЛЬНО!")
        else:
            print(f"❌ ФИЛЬТРАЦИЯ РАБОТАЕТ НЕПРАВИЛЬНО!")
        
        return success
        
    except Exception as e:
        print(f"💥 ОШИБКА В ТЕСТЕ: {e}")
        import traceback
        traceback.print_exc()
        return False

async def test_full_pipeline_with_filtering():
    """Тестируем полный pipeline с фильтрацией"""
    
    print("\n" + "="*60)
    print("🔍 ТЕСТ ПОЛНОГО PIPELINE С ФИЛЬТРАЦИЕЙ")
    print("="*60)
    
    # Тестовый запрос
    test_query = "Анкер забиваемый латунный М10\nБолт DIN 933 кл.пр.8.8 М10х30, цинк"
    
    try:
        from unittest.mock import Mock
        mock_message = Mock()
        mock_message.text = test_query
        mock_message.from_user.id = 12345
        mock_message.chat.id = 12345
        
        # Создаем pipeline
        pipeline = MessagePipeline(bot=None)
        
        print(f"📝 Тестовый запрос: '{test_query}'")
        
        # Запускаем полный pipeline
        result = await pipeline.process_message(mock_message, None)
        
        if result and result.get('excel_file'):
            print(f"✅ Pipeline завершен успешно")
            print(f"  - Excel файл: {result.get('excel_file')}")
            print(f"  - Статус: {result.get('status', 'N/A')}")
            
            # Проверяем содержимое Excel
            from openpyxl import load_workbook
            wb = load_workbook(result['excel_file'])
            ws = wb.active
            
            print(f"  - Строк в Excel: {ws.max_row}")
            print(f"  - Колонок в Excel: {ws.max_column}")
            
            # Показываем данные с вероятностями
            print(f"\n📊 ДАННЫЕ В EXCEL (только с вероятностью >= 60%):")
            for row in range(2, ws.max_row + 1):
                sku = ws.cell(row=row, column=11).value  # SKU
                probability = ws.cell(row=row, column=13).value  # Вероятность
                name = ws.cell(row=row, column=12).value  # Название
                print(f"  {row-1}. {sku} - {probability}% - {name}")
            
            return True
        else:
            print(f"❌ Pipeline не завершился успешно")
            return False
        
    except Exception as e:
        print(f"💥 ОШИБКА В ТЕСТЕ PIPELINE: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Запуск теста фильтрации по вероятности...")
    
    # Запускаем асинхронные тесты
    async def run_tests():
        # Тест фильтрации
        filter_test = await test_probability_filtering()
        
        # Тест полного pipeline
        pipeline_test = await test_full_pipeline_with_filtering()
        
        return filter_test and pipeline_test
    
    result = asyncio.run(run_tests())
    
    if result:
        print("\n🎉 ВСЕ ТЕСТЫ ФИЛЬТРАЦИИ ПРОЙДЕНЫ!")
    else:
        print("\n❌ ЕСТЬ ПРОБЛЕМЫ В ТЕСТАХ ФИЛЬТРАЦИИ!")
    
    sys.exit(0 if result else 1)
