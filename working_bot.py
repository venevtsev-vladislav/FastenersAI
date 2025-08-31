#!/usr/bin/env python3
"""
Рабочий Telegram Bot для поиска крепежных деталей
"""

import logging
import os
import asyncio
from datetime import datetime
from pathlib import Path
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from config import TELEGRAM_TOKEN
from utils.logger import setup_logging

# Настройка логирования
setup_logging()

# Дополнительное логирование в терминал
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)

logger = logging.getLogger(__name__)
logger.addHandler(console_handler)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    logger.info(f"Пользователь {update.effective_user.id} запустил бота")
    await update.message.reply_text(
        'Привет! Я бот для поиска крепежных деталей.\n\n'
        'Что я умею:\n'
        '• Искать детали по текстовому описанию\n'
        '• Обрабатывать загруженные файлы\n'
        '• Генерировать Excel отчеты\n\n'
        'Отправьте мне описание нужной детали или файл для анализа!'
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help"""
    logger.info(f"Пользователь {update.effective_user.id} запросил справку")
    await update.message.reply_text(
        '📋 **Как пользоваться ботом:**\n\n'
        '🔍 **Поиск по тексту:**\n'
        'Отправьте описание детали, например:\n'
        '• "болт М8х20"\n'
        '• "гайка шестигранная М10"\n'
        '• "шайба пружинная"\n\n'
        '📁 **Загрузка файлов:**\n'
        '• Excel файлы (.xlsx, .xls)\n'
        '• Текстовые файлы (.txt)\n'
        '• Изображения деталей\n\n'
        '📊 **Результаты:**\n'
        'Бот найдет подходящие детали и сгенерирует Excel отчет.'
    )

async def generate_excel_report(query: str, results: list, user_id: int) -> str:
    """Генерирует Excel отчет с результатами поиска"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Создаем папку для отчетов, если её нет
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        
        # Формируем данные для Excel
        excel_data = []
        for i, part in enumerate(results, 1):
            excel_data.append({
                '№': i,
                'Запрос от пользователя': query,
                'Кол-во запрашиваемых деталей': part.get('quantity', 1),
                'Артикул (SKU)': part.get('sku', 'N/A'),
                'Наименование в каталоге': part.get('name', 'N/A')
            })
        
        # Создаем DataFrame
        df = pd.DataFrame(excel_data)
        
        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{user_id}_{timestamp}.xlsx"
        filepath = reports_dir / filename
        
        # Создаем Excel файл с форматированием
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Результаты поиска', index=False)
            
            # Получаем рабочую книгу для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Результаты поиска']
            
            # Форматируем заголовки
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Автоматически подгоняем ширину столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Excel отчет создан: {filepath}")
        return str(filepath)
        
    except ImportError:
        logger.error("Библиотеки pandas или openpyxl не установлены")
        return None
    except Exception as e:
        logger.error(f"Ошибка при создании Excel отчета: {e}")
        return None

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загруженных документов"""
    document = update.message.document
    file_name = document.file_name
    file_size = document.file_size
    user_id = update.effective_user.id
    
    logger.info(f"Пользователь {user_id} загрузил документ: {file_name} ({file_size} байт)")
    
    await update.message.reply_text(
        f'📁 **Получен документ:**\n'
        f'Название: {file_name}\n'
        f'Размер: {file_size} байт\n\n'
        f'⏳ Обрабатываю файл...'
    )
    
    # Имитируем обработку
    await asyncio.sleep(3)
    
    # Определяем тип файла
    if file_name.lower().endswith(('.xlsx', '.xls')):
        await update.message.reply_text(
            f'✅ **Excel файл обработан!**\n\n'
            f'Файл: {file_name}\n'
            f'📊 Анализирую содержимое...\n\n'
            f'🔍 Ищу подходящие детали...'
        )
        
        # Имитируем поиск
        await asyncio.sleep(2)
        
        # Создаем имитационные результаты
        mock_results = [
            {'sku': '5-0011450', 'name': 'Анкерный болт 8х100', 'quantity': 15},
            {'sku': '5-0011480', 'name': 'Анкерный болт 8х80', 'quantity': 12},
            {'sku': '5-0011490', 'name': 'Анкерный болт 8х90', 'quantity': 8}
        ]
        
        await update.message.reply_text(
            f'🎯 **Результаты поиска:**\n\n'
            f'Найдено: **{len(mock_results)} деталей**\n'
            f'📋 Генерирую Excel отчет...'
        )
        
        # Генерируем Excel отчет
        excel_file = await generate_excel_report(f"Анализ файла: {file_name}", mock_results, user_id)
        
        if excel_file:
            # Отправляем файл пользователю
            with open(excel_file, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"report_{file_name.replace('.xlsx', '').replace('.xls', '')}.xlsx",
                    caption=f'📊 **Отчет готов!**\n\n'
                            f'Файл: `report_{file_name}`\n'
                            f'Строк: {len(mock_results)}\n'
                            f'Колонок: 5\n\n'
                            f'⚠️ **Примечание:** Функция Excel пока в разработке.\n'
                            f'Реальные данные будут доступны после настройки Supabase.'
                )
            
            # Удаляем временный файл
            try:
                os.remove(excel_file)
                logger.info(f"Временный файл удален: {excel_file}")
            except:
                pass
        else:
            await update.message.reply_text('❌ Ошибка при создании Excel отчета')
        
    elif file_name.lower().endswith('.txt'):
        await update.message.reply_text(
            f'📝 **Текстовый файл получен!**\n\n'
            f'Файл: {file_name}\n'
            f'📖 Анализирую содержимое...\n\n'
            f'🔍 Ищу упоминания деталей...'
        )
        
        await asyncio.sleep(2)
        
        await update.message.reply_text(
            f'📋 **Анализ завершен:**\n\n'
            f'Найдено упоминаний деталей: **7**\n'
            f'Основные категории:\n'
            f'• Болты и винты\n'
            f'• Гайки и шайбы\n'
            f'• Анкеры и дюбели\n\n'
            f'⚠️ **Примечание:** Функция анализа пока в разработке.'
        )
        
    else:
        await update.message.reply_text(
            f'❓ **Неизвестный тип файла:**\n\n'
            f'Файл: {file_name}\n'
            f'Поддерживаемые форматы:\n'
            f'• Excel: .xlsx, .xls\n'
            f'• Текст: .txt\n'
            f'• Изображения: .jpg, .png\n\n'
            f'Попробуйте загрузить файл другого формата.'
        )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загруженных фотографий"""
    photo = update.message.photo[-1]
    file_size = photo.file_size
    user_id = update.effective_user.id
    
    logger.info(f"Пользователь {user_id} загрузил фото: {file_size} байт")
    
    await update.message.reply_text(
        f'📸 **Получено изображение!**\n\n'
        f'Размер: {file_size} байт\n'
        f'⏳ Анализирую изображение...'
    )
    
    # Имитируем анализ изображения
    await asyncio.sleep(3)
    
    await update.message.reply_text(
        f'🔍 **Анализ изображения:**\n\n'
        f'✅ Изображение обработано\n'
        f'📋 Определяю тип детали...\n\n'
        f'🎯 **Результат:**\n'
        f'Тип: Крепежная деталь\n'
        f'Категория: Болт/Винт\n'
        f'Размер: Приблизительно М8-М12\n\n'
        f'⚠️ **Примечание:** Функция распознавания изображений пока в разработке.\n'
        f'Реальные результаты будут доступны после настройки AI.'
    )

async def echo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик всех текстовых сообщений"""
    if update.message.text:
        text = update.message.text
        user_id = update.effective_user.id
        
        logger.info(f"Пользователь {user_id} отправил запрос: {text}")
        
        await update.message.reply_text(
            f'🔍 **Поиск деталей:**\n\n'
            f'Запрос: "{text}"\n'
            f'⏳ Ищу подходящие детали...'
        )
        
        # Имитируем поиск
        await asyncio.sleep(2)
        
        # Создаем имитационные результаты
        mock_results = [
            {'sku': '5-0011450', 'name': 'Анкерный болт 8х100', 'quantity': 12},
            {'sku': '5-0011480', 'name': 'Анкерный болт 8х80', 'quantity': 8},
            {'sku': '5-0011490', 'name': 'Анкерный болт 8х90', 'quantity': 15}
        ]
        
        await update.message.reply_text(
            f'🎯 **Результаты поиска:**\n\n'
            f'Запрос: "{text}"\n'
            f'Найдено: **{len(mock_results)} деталей**\n\n'
            f'📋 **Топ результаты:**\n'
            f'1. {mock_results[0]["name"]} (SKU: {mock_results[0]["sku"]})\n'
            f'2. {mock_results[1]["name"]} (SKU: {mock_results[1]["sku"]})\n'
            f'3. {mock_results[2]["name"]} (SKU: {mock_results[2]["sku"]})\n\n'
            f'📊 Генерирую Excel отчет...'
        )
        
        # Генерируем Excel отчет
        excel_file = await generate_excel_report(text, mock_results, user_id)
        
        if excel_file:
            # Отправляем файл пользователю
            with open(excel_file, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"search_results_{text.replace(' ', '_')}.xlsx",
                    caption=f'📊 **Отчет готов!**\n\n'
                            f'Файл: `search_results_{text.replace(" ", "_")}.xlsx`\n'
                            f'Результатов: {len(mock_results)}\n'
                            f'Колонок: 5\n\n'
                            f'⚠️ **Примечание:** Функция Excel пока в разработке.\n'
                            f'Реальные данные будут доступны после настройки Supabase.'
                )
            
            # Удаляем временный файл
            try:
                os.remove(excel_file)
                logger.info(f"Временный файл удален: {excel_file}")
            except:
                pass
        else:
            await update.message.reply_text('❌ Ошибка при создании Excel отчета')

def main():
    """Основная функция"""
    logger.info("Запускаю бота...")
    
    # Создаем приложение
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Добавляем обработчики команд
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    
    # Добавляем обработчики файлов
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    
    # Добавляем обработчик текстовых сообщений (последним)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, echo))
    
    logger.info("Бот запущен!")
    
    # Запускаем
    app.run_polling()

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем")
    except Exception as e:
        logger.error(f"Ошибка: {e}")
