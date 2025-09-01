"""
Тест для проверки исправления bug с процентами в Excel
"""

import pytest
import tempfile
import os
from openpyxl import load_workbook
from services.excel_generator import ExcelGenerator


class TestExcelPercentBug:
    """Тесты для проверки исправления bug с процентами"""
    
    def setup_method(self):
        """Настройка перед каждым тестом"""
        self.generator = ExcelGenerator()
        self.test_results = [
            {
                'order_position': 1,
                'full_query': 'болт M6x40',
                'search_query': 'болт M6x40',
                'requested_quantity': 10,
                'sku': 'B001',
                'name': 'Болт M6x40 DIN 933',
                'type': 'болт',
                'pack_size': 100,
                'unit': 'шт',
                'relevance_score': 95,
                'probability_percent': 95,
                'validation_status': 'APPROVED',
                'is_normalized': False,
                'improvement_cycle': 0,
                'is_normalized': False
            },
            {
                'order_position': 2,
                'full_query': 'гайка M6',
                'search_query': 'гайка M6',
                'requested_quantity': 5,
                'sku': 'N001',
                'name': 'Гайка M6 DIN 934',
                'type': 'гайка',
                'pack_size': 50,
                'unit': 'шт',
                'relevance_score': 70,
                'probability_percent': 70,
                'validation_status': 'APPROVED',
                'is_normalized': False,
                'improvement_cycle': 0,
                'is_normalized': False
            }
        ]
    
    @pytest.mark.asyncio
    async def test_excel_percent_formatting(self):
        """Тест правильного форматирования процентов в Excel"""
        # Генерируем Excel файл
        excel_file = await self.generator.generate_excel(
            search_results=self.test_results,
            user_request="болт M6x40, гайка M6"
        )
        
        try:
            # Загружаем файл для проверки
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Проверяем заголовки
            assert worksheet.cell(row=1, column=11).value == "Релевантность"
            assert worksheet.cell(row=1, column=12).value == "Вероятность (%)"
            
            # Проверяем данные
            # Релевантность (колонка 11) - должно быть число от 0 до 1
            relevance_cell = worksheet.cell(row=2, column=11)
            assert relevance_cell.value == 0.95  # 95% -> 0.95
            assert relevance_cell.number_format == '0.00'
            
            relevance_cell2 = worksheet.cell(row=3, column=11)
            assert relevance_cell2.value == 0.70  # 70% -> 0.70
            assert relevance_cell2.number_format == '0.00'
            
            # Вероятность (колонка 12) - должно быть число от 0 до 1 с форматом процентов
            probability_cell = worksheet.cell(row=2, column=12)
            assert probability_cell.value == 0.95  # 95% -> 0.95
            assert probability_cell.number_format == '0.00%'
            
            probability_cell2 = worksheet.cell(row=3, column=12)
            assert probability_cell2.value == 0.70  # 70% -> 0.70
            assert probability_cell2.number_format == '0.00%'
            
            # Проверяем, что значения не равны 1.0 (старый bug)
            assert relevance_cell.value != 1.0
            assert probability_cell.value != 1.0
            
        finally:
            # Очищаем временный файл
            if os.path.exists(excel_file):
                os.unlink(excel_file)
    
    @pytest.mark.asyncio
    async def test_excel_data_completeness(self):
        """Тест полноты данных в Excel"""
        excel_file = await self.generator.generate_excel(
            search_results=self.test_results,
            user_request="тестовый запрос"
        )
        
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Проверяем количество строк (заголовок + данные)
            assert worksheet.max_row == 3  # 1 заголовок + 2 результата
            
            # Проверяем количество колонок
            assert worksheet.max_column == 17
            
            # Проверяем конкретные данные
            # Позиция в заказе
            assert worksheet.cell(row=2, column=2).value == 1
            assert worksheet.cell(row=3, column=2).value == 2
            
            # SKU
            assert worksheet.cell(row=2, column=6).value == 'B001'
            assert worksheet.cell(row=3, column=6).value == 'N001'
            
            # Название
            assert 'Болт M6x40 DIN 933' in str(worksheet.cell(row=2, column=7).value)
            assert 'Гайка M6 DIN 934' in str(worksheet.cell(row=3, column=7).value)
            
        finally:
            if os.path.exists(excel_file):
                os.unlink(excel_file)
    
    @pytest.mark.asyncio
    async def test_excel_edge_cases(self):
        """Тест граничных случаев"""
        # Пустые результаты
        empty_results = []
        excel_file = await self.generator.generate_excel(
            search_results=empty_results,
            user_request="пустой запрос"
        )
        
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Должен быть только заголовок
            assert worksheet.max_row == 1
            assert worksheet.max_column == 17
            
        finally:
            if os.path.exists(excel_file):
                os.unlink(excel_file)
        
        # Результаты с отсутствующими полями
        incomplete_results = [
            {
                'sku': 'TEST001',
                'name': 'Тестовая деталь'
                # Остальные поля отсутствуют
            }
        ]
        
        excel_file = await self.generator.generate_excel(
            search_results=incomplete_results,
            user_request="неполный запрос"
        )
        
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Проверяем, что файл создался без ошибок
            assert worksheet.max_row == 2  # 1 заголовок + 1 результат
            assert worksheet.max_column == 17
            
            # Проверяем, что отсутствующие поля заполнены пустыми значениями
            assert worksheet.cell(row=2, column=2).value is None  # order_position
            assert worksheet.cell(row=2, column=6).value == 'TEST001'  # sku
            assert worksheet.cell(row=2, column=7).value == 'Тестовая деталь'  # name
            
        finally:
            if os.path.exists(excel_file):
                os.unlink(excel_file)


if __name__ == '__main__':
    pytest.main([__file__])
