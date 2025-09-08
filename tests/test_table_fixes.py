import asyncio
import os
import sys
from openpyxl import load_workbook

# Ensure project root is on the Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Set required environment variables for imports
os.environ.setdefault('TELEGRAM_BOT_TOKEN', 'test')
os.environ.setdefault('OPENAI_API_KEY', 'test')
os.environ.setdefault('SUPABASE_URL', 'http://localhost')
os.environ.setdefault('SUPABASE_KEY', 'test')

from services.excel_generator_v2 import ExcelGeneratorV2
from pipeline.text_parser import TextParser
from pipeline.processing_pipeline import ProcessingResult
from pipeline.matching_engine import MatchCandidate

def test_text_parser_preserves_commas():
    parser = TextParser()
    lines = parser.parse_text_input("Гвоздь строительный 4,0х120")
    assert len(lines) == 1
    assert "4,0х120" in lines[0].raw_text

def test_excel_summary_headers_and_price_blank():
    generator = ExcelGeneratorV2()
    result = ProcessingResult(
        line_id=1,
        raw_text='Гвоздь строительный 4,0х120',
        normalized_text='Гвоздь строительный 4,0х120',
        chosen_ku='SKU123',
        qty_packs=1,
        qty_units=100,
        unit='шт',
        price=None,
        total=None,
        status='ok',
        chosen_method='rules',
        candidates=[
            MatchCandidate(
                ku='SKU123',
                name='Гвоздь строительный 4,0х120',
                pack_qty=100,
                price=None,
                unit='шт',
                score=1.0,
                explanation='',
                source='rules'
            )
        ]
    )
    request_data = {'input_lines': [result.raw_text], 'source': 'text'}
    file_path = asyncio.run(
        generator.generate_excel('test', [result], request_data)
    )
    wb = load_workbook(file_path)
    sheet = wb['Итог']
    assert sheet.cell(1,3).value == 'SKU'
    assert sheet.cell(1,6).value == 'Единица_изм'
    assert sheet.cell(2,4).value == 'Гвоздь строительный 4,0х120'
    assert sheet.cell(2,5).value == 100
    assert sheet.cell(2,6).value == 'шт'
    assert sheet.cell(2,9).value in (None, '')
