#!/usr/bin/env python3
"""
Полный тест флоу для запроса "Саморез 4,2-25"
Проходит весь pipeline: GPT парсинг → Supabase поиск → генерация Excel
"""

import asyncio
import sys
import os
import json
from datetime import datetime

# Добавляем родительскую папку в путь для импорта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from pipeline.message_processor import MessagePipeline
from services.excel_generator import ExcelGenerator

async def test_full_flow():
    """Тестируем полный флоу для запроса 'Саморез 4,2-25'"""
    
    print("🔍 ПОЛНЫЙ ТЕСТ ФЛОУ: 'Саморез 4,2-25'")
    print("=" * 60)
    
    # Тестовый запрос
    test_query = "Саморез 4,2-25"
    user_id = 12345
    
    try:
        # Создаем pipeline (без bot для тестирования)
        pipeline = MessagePipeline(bot=None)
        
        print("✅ Pipeline создан успешно")
        print(f"📝 Тестовый запрос: '{test_query}'")
        print(f"👤 User ID: {user_id}")
        
        # ЭТАП 1: Парсинг и нормализация
        print("\n" + "="*60)
        print("🔍 ЭТАП 1: ПАРСИНГ И НОРМАЛИЗАЛИЗАЦИЯ")
        print("="*60)
        
        # Создаем mock Message объект
        from unittest.mock import Mock
        mock_message = Mock()
        mock_message.text = test_query
        mock_message.from_user.id = user_id
        mock_message.chat.id = user_id
        
        normalized_result = await pipeline._parse_and_normalize(mock_message)
        
        print(f"✅ Результат парсинга:")
        print(f"  - Обработанный текст: {normalized_result.get('processed_text', 'N/A')}")
        print(f"  - User intent: {json.dumps(normalized_result.get('user_intent', {}), ensure_ascii=False, indent=2)}")
        print(f"  - Confidence: {normalized_result.get('confidence', 'N/A')}")
        
        # ЭТАП 2: Поиск в базе данных
        print("\n" + "="*60)
        print("🔍 ЭТАП 2: ПОИСК В БАЗЕ ДАННЫХ")
        print("="*60)
        
        search_results = await pipeline._search_database(normalized_result)
        
        print(f"✅ Результат поиска:")
        print(f"  - Найдено результатов: {len(search_results)}")
        
        if search_results:
            print(f"  - Первый результат:")
            first_result = search_results[0]
            for key, value in first_result.items():
                print(f"    {key}: {value}")
        else:
            print("  ❌ Результаты не найдены!")
            return False
        
        # ЭТАП 3: Ранжирование результатов
        print("\n" + "="*60)
        print("🔍 ЭТАП 3: РАНЖИРОВАНИЕ РЕЗУЛЬТАТОВ")
        print("="*60)
        
        ranked_results = await pipeline._rank_results(search_results, normalized_result)
        
        print(f"✅ Результат ранжирования:")
        print(f"  - Ранжировано результатов: {len(ranked_results)}")
        
        # ЭТАП 4: Генерация Excel
        print("\n" + "="*60)
        print("🔍 ЭТАП 4: ГЕНЕРАЦИЯ EXCEL")
        print("="*60)
        
        excel_file = await pipeline._generate_excel(ranked_results, normalized_result)
        
        print(f"✅ Excel файл создан: {excel_file}")
        print(f"📁 Размер файла: {os.path.getsize(excel_file)} байт")
        
        # Проверяем содержимое Excel файла
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📋 СОДЕРЖИМОЕ EXCEL ФАЙЛА:")
        print(f"  - Лист: {ws.title}")
        print(f"  - Строк: {ws.max_row}")
        print(f"  - Колонок: {ws.max_column}")
        
        # Показываем заголовки
        print(f"\n📝 ЗАГОЛОВКИ:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"  {col:2d}: {header}")
        
        # Показываем данные
        print(f"\n📊 ДАННЫЕ:")
        for row in range(2, ws.max_row + 1):
            print(f"  Строка {row}:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                print(f"    {col:2d}: {value}")
        
        # ЭТАП 5: Финальная обработка
        print("\n" + "="*60)
        print("🔍 ЭТАП 5: ФИНАЛЬНАЯ ОБРАБОТКА")
        print("="*60)
        
        final_result = await pipeline._finalize_results(ranked_results, excel_file, normalized_result)
        
        print(f"✅ Финальный результат:")
        print(f"  - Excel файл: {final_result.get('excel_file', 'N/A')}")
        print(f"  - Количество результатов: {final_result.get('results_count', 'N/A')}")
        print(f"  - Статус: {final_result.get('status', 'N/A')}")
        
        print("\n" + "="*60)
        print("🎉 ПОЛНЫЙ ФЛОУ ЗАВЕРШЕН УСПЕШНО!")
        print("="*60)
        
        return True
        
    except Exception as e:
        print(f"\n💥 ОШИБКА В ФЛОУ: {e}")
        import traceback
        traceback.print_exc()
        return False

async def test_excel_generator_directly():
    """Тестируем генератор Excel напрямую с тестовыми данными"""
    
    print("\n" + "="*60)
    print("🔍 ДОПОЛНИТЕЛЬНЫЙ ТЕСТ: ГЕНЕРАТОР EXCEL НАПРЯМУЮ")
    print("="*60)
    
    # Тестовые данные, имитирующие результат поиска
    test_search_results = [
        {
            "order_position": 1,
            "search_query": "саморез 4,2-25",
            "diameter": "4.2",
            "length": "25",
            "material": "сталь",
            "coating": "оцинкованный",
            "requested_quantity": 100,
            "confidence": 0.95,
            "sku": "10-0010040",
            "name": "Саморез ШСММ АРТ 9999 A2 4,2X25",
            "smart_probability": 0.92,
            "pack_size": 300,
            "unit": "шт"
        }
    ]
    
    user_request = "Саморез 4,2-25"
    
    try:
        # Создаем генератор Excel
        excel_generator = ExcelGenerator()
        
        print("✅ Excel генератор создан")
        print(f"📊 Тестовые данные: {len(test_search_results)} позиций")
        
        # Генерируем Excel файл
        excel_file = await excel_generator.generate_excel(
            search_results=test_search_results,
            user_request=user_request
        )
        
        print(f"✅ Excel файл создан: {excel_file}")
        print(f"📁 Размер файла: {os.path.getsize(excel_file)} байт")
        
        # Проверяем содержимое файла
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📋 СОДЕРЖИМОЕ EXCEL ФАЙЛА:")
        print(f"  - Лист: {ws.title}")
        print(f"  - Строк: {ws.max_row}")
        print(f"  - Колонок: {ws.max_column}")
        
        # Показываем заголовки
        print(f"\n📝 ЗАГОЛОВКИ:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"  {col:2d}: {header}")
        
        # Показываем данные
        print(f"\n📊 ДАННЫЕ:")
        for row in range(2, ws.max_row + 1):
            print(f"  Строка {row}:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                print(f"    {col:2d}: {value}")
        
        return True
        
    except Exception as e:
        print(f"💥 ОШИБКА при генерации Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Запуск полного теста флоу...")
    
    # Запускаем асинхронные тесты
    async def run_tests():
        # Сначала тестируем генератор напрямую
        direct_test = await test_excel_generator_directly()
        
        # Затем полный флоу
        full_test = await test_full_flow()
        
        return direct_test and full_test
    
    result = asyncio.run(run_tests())
    
    if result:
        print("\n🎉 ВСЕ ТЕСТЫ ПРОЙДЕНЫ УСПЕШНО!")
    else:
        print("\n❌ ЕСТЬ ПРОБЛЕМЫ В ТЕСТАХ!")
    
    sys.exit(0 if result else 1)
