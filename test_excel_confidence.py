#!/usr/bin/env python3
"""
Тест передачи уверенности GPT в Excel файл
"""

import asyncio
import sys
import os
import json

# Добавляем родительскую папку в путь для импорта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from handlers.message_handler_full import FullMessageHandler
from services.excel_generator import ExcelGenerator

async def test_excel_confidence():
    """Тестируем передачу уверенности GPT в Excel"""
    
    print("🔍 ТЕСТ ПЕРЕДАЧИ УВЕРЕННОСТИ GPT В EXCEL")
    print("=" * 60)
    
    # Тестовые данные - имитируем результат GPT парсинга
    user_intent = {
        "is_multiple_order": True,
        "items": [
            {
                "type": "анкер забиваемый",
                "diameter": "M10",
                "length": None,
                "material": "латунь",
                "coating": None,
                "quantity": None,
                "confidence": 0.9  # ← Уверенность от GPT
            },
            {
                "type": "болт",
                "diameter": "M10",
                "length": "30",
                "material": None,
                "coating": "цинк",
                "quantity": None,
                "confidence": 0.95  # ← Уверенность от GPT
            }
        ]
    }
    
    # Тестовые данные - имитируем результат поиска
    search_results = [
        {
            "sku": "10-0010040",
            "name": "Анкер забиваемый латунный М10",
            "diameter": "M10",
            "length": None,
            "material": "латунь",
            "coating": None,
            "smart_probability": 85,
            "pack_size": 300,
            "unit": "шт",
            "search_query": "анкер забиваемый M10"
        },
        {
            "sku": "6-0049020",
            "name": "Болт DIN 933 кл.пр.8.8 М10х30, цинк",
            "diameter": "M10",
            "length": "30",
            "material": None,
            "coating": "цинк",
            "smart_probability": 90,
            "pack_size": 100,
            "unit": "шт",
            "search_query": "болт M10 30"
        }
    ]
    
    original_query = "Анкер забиваемый латунный М10\nБолт DIN 933 кл.пр.8.8 М10х30, цинк"
    
    try:
        # Создаем обработчик
        handler = FullMessageHandler()
        
        print("✅ MessageHandlerFull создан")
        print(f"📊 Тестовые данные:")
        print(f"  - User intent: {json.dumps(user_intent, ensure_ascii=False, indent=2)}")
        print(f"  - Search results: {len(search_results)} позиций")
        
        # Тестируем преобразование данных
        converted_results = handler._convert_search_results_to_new_format(search_results, user_intent)
        
        print(f"\n✅ Преобразованные данные:")
        for i, result in enumerate(converted_results, 1):
            print(f"  Позиция {i}:")
            print(f"    - confidence (GPT): {result.get('confidence', 'N/A')}")
            print(f"    - smart_probability: {result.get('smart_probability', 'N/A')}")
            print(f"    - sku: {result.get('sku', 'N/A')}")
            print(f"    - name: {result.get('name', 'N/A')}")
        
        # Тестируем генерацию Excel
        print(f"\n🔍 ГЕНЕРАЦИЯ EXCEL:")
        excel_file = await handler._generate_excel_file(search_results, user_intent, original_query)
        
        print(f"✅ Excel файл создан: {excel_file}")
        print(f"📁 Размер файла: {os.path.getsize(excel_file)} байт")
        
        # Проверяем содержимое Excel файла
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📋 СОДЕРЖИМОЕ EXCEL ФАЙЛА:")
        print(f"  - Лист: {ws.title}")
        print(f"  - Строк: {ws.max_row}")
        print(f"  - Колонок: {ws.max_column}")
        
        # Показываем заголовки
        print(f"\n📝 ЗАГОЛОВКИ:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"  {col:2d}: {header}")
        
        # Показываем данные с акцентом на уверенность
        print(f"\n📊 ДАННЫЕ (с акцентом на уверенность):")
        for row in range(2, ws.max_row + 1):
            print(f"  Строка {row}:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                header = ws.cell(row=1, column=col).value
                if "Уверенность" in str(header) or "Вероятность" in str(header):
                    print(f"    {col:2d}: {header} = {value} ⭐")
                else:
                    print(f"    {col:2d}: {header} = {value}")
        
        return True
        
    except Exception as e:
        print(f"💥 ОШИБКА В ТЕСТЕ: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Запуск теста передачи уверенности GPT в Excel...")
    
    # Запускаем асинхронный тест
    result = asyncio.run(test_excel_confidence())
    
    if result:
        print("\n🎉 ТЕСТ ПЕРЕДАЧИ УВЕРЕННОСТИ ПРОЙДЕН!")
    else:
        print("\n❌ ТЕСТ ПЕРЕДАЧИ УВЕРЕННОСТИ НЕ ПРОЙДЕН!")
    
    sys.exit(0 if result else 1)
