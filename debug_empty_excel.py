#!/usr/bin/env python3
"""
Отладка пустого Excel файла
"""

import asyncio
import sys
import os
import json

# Добавляем родительскую папку в путь для импорта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from pipeline.message_processor import MessagePipeline

async def debug_empty_excel():
    """Отладка пустого Excel файла"""
    
    print("🔍 ОТЛАДКА ПУСТОГО EXCEL ФАЙЛА")
    print("=" * 60)
    
    # Тестовый запрос
    test_query = "Анкер забиваемый латунный М10"
    
    try:
        from unittest.mock import Mock
        mock_message = Mock()
        mock_message.text = test_query
        mock_message.from_user.id = 12345
        mock_message.chat.id = 12345
        
        # Создаем pipeline
        pipeline = MessagePipeline(bot=None)
        
        print(f"📝 Тестовый запрос: '{test_query}'")
        
        # ЭТАП 1: Парсинг
        print("\n🔍 ЭТАП 1: ПАРСИНГ")
        normalized_result = await pipeline._parse_and_normalize(mock_message)
        
        if normalized_result:
            print(f"✅ Парсинг успешен")
            print(f"  - User intent: {json.dumps(normalized_result.get('user_intent', {}), ensure_ascii=False, indent=2)}")
        else:
            print(f"❌ Парсинг не удался")
            return False
        
        # ЭТАП 2: Поиск
        print("\n🔍 ЭТАП 2: ПОИСК")
        search_results = await pipeline._search_database(normalized_result)
        
        print(f"✅ Поиск завершен: {len(search_results)} результатов")
        if search_results:
            print(f"  - Первый результат:")
            first_result = search_results[0]
            for key, value in first_result.items():
                if key in ['sku', 'name', 'smart_probability', 'probability_percent', 'relevance_score']:
                    print(f"    {key}: {value}")
        else:
            print(f"❌ Результаты поиска не найдены")
            return False
        
        # ЭТАП 3: Ранжирование и фильтрация
        print("\n🔍 ЭТАП 3: РАНЖИРОВАНИЕ И ФИЛЬТРАЦИЯ")
        ranked_results = await pipeline._rank_results(search_results, normalized_result)
        
        print(f"✅ Ранжирование завершено: {len(ranked_results)} результатов")
        if ranked_results:
            print(f"  - Результаты после фильтрации:")
            for i, result in enumerate(ranked_results, 1):
                sku = result.get('sku', 'N/A')
                probability = result.get('smart_probability', 0)
                name = result.get('name', 'N/A')
                print(f"    {i}. {sku} - {probability}% - {name}")
        else:
            print(f"❌ Все результаты отфильтрованы!")
            
            # Показываем, что было отфильтровано
            print(f"  - Исходные результаты:")
            for i, result in enumerate(search_results, 1):
                sku = result.get('sku', 'N/A')
                probability = result.get('smart_probability', 0)
                name = result.get('name', 'N/A')
                print(f"    {i}. {sku} - {probability}% - {name}")
        
        # ЭТАП 4: Генерация Excel
        print("\n🔍 ЭТАП 4: ГЕНЕРАЦИЯ EXCEL")
        if ranked_results:
            excel_file = await pipeline._generate_excel(ranked_results, normalized_result)
            print(f"✅ Excel файл создан: {excel_file}")
            
            # Проверяем содержимое
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            print(f"  - Строк в Excel: {ws.max_row}")
            print(f"  - Колонок в Excel: {ws.max_column}")
            
            if ws.max_row > 1:
                print(f"✅ Excel содержит данные")
            else:
                print(f"❌ Excel пустой (только заголовки)")
        else:
            print(f"❌ Excel не создан - нет данных для генерации")
        
        return len(ranked_results) > 0
        
    except Exception as e:
        print(f"💥 ОШИБКА В ОТЛАДКЕ: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Запуск отладки пустого Excel...")
    
    result = asyncio.run(debug_empty_excel())
    
    if result:
        print("\n🎉 ПРОБЛЕМА НЕ НАЙДЕНА!")
    else:
        print("\n❌ НАЙДЕНА ПРОБЛЕМА - ВСЕ РЕЗУЛЬТАТЫ ОТФИЛЬТРОВАНЫ!")
    
    sys.exit(0 if result else 1)
