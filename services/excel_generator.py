"""
Сервис для генерации Excel файлов с результатами поиска
"""

import logging
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import MAX_EXCEL_ROWS

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Класс для генерации Excel файлов с результатами поиска"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    async def generate_excel(self, search_results: list, user_request: str) -> str:
        """Генерирует Excel файл с результатами поиска"""
        try:
            # Сохраняем запрос пользователя для использования в заполнении данных
            self.user_request = user_request
            
            # Ограничиваем количество строк
            if len(search_results) > MAX_EXCEL_ROWS:
                search_results = search_results[:MAX_EXCEL_ROWS]
                logger.warning(f"Ограничили результаты до {MAX_EXCEL_ROWS} строк")
            
            # Создаем рабочую книгу
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Результаты поиска"
            
            # Настраиваем заголовки
            self._setup_headers()
            
            # Заполняем данными
            self._fill_data(search_results)
            
            # Настраиваем стили
            self._apply_styles()
            
            # Автоматически подгоняем ширину колонок
            self._auto_adjust_columns()
            
            # Создаем временный файл
            temp_file = self._create_temp_file()
            
            # Сохраняем
            self.workbook.save(temp_file)
            logger.info(f"Excel файл создан: {temp_file}")
            
            return temp_file
            
        except Exception as e:
            logger.error(f"Ошибка при создании Excel файла: {e}")
            raise
    
    def _setup_headers(self):
        """Настраивает заголовки таблицы"""
        headers = [
            "№ п/п",
            "№ позиции в запросе клиента",
            "Запрос пользователя",
            "Поисковый запрос",
            "Диаметр",
            "Длина",
            "Материал",
            "Покрытие",
            "Количество",
            "Уверенность GPT",
            "Найденный SKU",
            "Наименование в каталоге",
            "Вероятность",
            "Размер упаковки",
            "Единица измерения",
            "Стандарт",
            "Класс прочности",
            "Статус"
        ]
        
        logger.info(f"Настраиваем заголовки: {len(headers)} колонок")
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            logger.debug(f"Заголовок {col}: {header}")
    
    def _fill_data(self, search_results: list):
        """Заполняет таблицу данными - ОБНОВЛЕНО для 18 колонок"""
        logger.info(f"Заполняем данные: {len(search_results)} результатов")
        
        if not search_results:
            logger.warning("Нет данных для заполнения Excel")
            return
            
        for row, item in enumerate(search_results, 2):
            logger.debug(f"Заполняем строку {row} данными: {item}")
            
            # Проверяем, что item - это словарь
            if not isinstance(item, dict):
                logger.error(f"Элемент {row-1} не является словарем: {type(item)} - {item}")
                continue
                
            # № п/п
            self.worksheet.cell(row=row, column=1, value=row - 1)
            
            # № позиции в запросе клиента
            order_position = item.get('order_position', item.get('original_position', ''))
            self.worksheet.cell(row=row, column=2, value=order_position)
            
            # Запрос пользователя (общий)
            self.worksheet.cell(row=row, column=3, value=self.user_request)
            
            # Поисковый запрос (для конкретной детали)
            search_query = item.get('search_query', '')
            self.worksheet.cell(row=row, column=4, value=search_query)
            
            # Диаметр
            diameter = item.get('diameter', '')
            self.worksheet.cell(row=row, column=5, value=diameter)
            
            # Длина
            length = item.get('length', '')
            self.worksheet.cell(row=row, column=6, value=length)
            
            # Материал
            material = item.get('material', '')
            self.worksheet.cell(row=row, column=7, value=material)
            
            # Покрытие (берем из user_intent или из названия)
            coating = item.get('coating', '')
            if not coating:
                # Пытаемся извлечь из названия
                name = item.get('name', '').lower()
                if 'оцинкованный' in name:
                    coating = 'оцинкованный'
                elif 'латунный' in name:
                    coating = 'латунный'
                elif 'нержавеющий' in name:
                    coating = 'нержавеющий'
            self.worksheet.cell(row=row, column=8, value=coating)
            
            # Количество
            requested_quantity = item.get('requested_quantity', item.get('quantity', 1))
            self.worksheet.cell(row=row, column=9, value=requested_quantity)
            
            # Уверенность GPT (берем из user_intent)
            confidence = item.get('confidence', 0)
            if not confidence:
                # Fallback на confidence_score из Supabase
                confidence = item.get('confidence_score', 0)
            confidence_cell = self.worksheet.cell(row=row, column=10, value=confidence)
            confidence_cell.number_format = '0.00'
            
            # Найденный SKU
            sku = item.get('sku', '')
            self.worksheet.cell(row=row, column=11, value=sku)
            
            # Наименование в каталоге
            name = item.get('name', '')
            self.worksheet.cell(row=row, column=12, value=name)
            
            # Вероятность (умная вероятность)
            smart_probability = item.get('smart_probability', 0)
            self.worksheet.cell(row=row, column=13, value=smart_probability)
            
            # Размер упаковки
            pack_size = item.get('pack_size', 0)
            self.worksheet.cell(row=row, column=14, value=pack_size)
            
            # Единица измерения
            unit = item.get('unit', 'шт')
            self.worksheet.cell(row=row, column=15, value=unit)
            
            # Стандарт
            standard = self._extract_standard(name)
            self.worksheet.cell(row=row, column=16, value=standard)
            
            # Класс прочности
            strength_class = self._extract_strength_class(name)
            self.worksheet.cell(row=row, column=17, value=strength_class)
            
            # Статус
            status = 'Найдено' if sku else 'Не найдено'
            self.worksheet.cell(row=row, column=18, value=status)
    
    def _apply_styles(self):
        """Применяет стили к таблице"""
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем границы ко всем ячейкам с данными
        for row in self.worksheet.iter_rows(min_row=1, max_row=self.worksheet.max_row, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Выравнивание заголовков по центру
        for col in range(1, 19):
            cell = self.worksheet.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _auto_adjust_columns(self):
        """Автоматически подгоняет ширину колонок"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_temp_file(self) -> str:
        """Создает временный файл для Excel"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.gettempdir()
        filename = f"search_results_{timestamp}.xlsx"
        return os.path.join(temp_dir, filename)
    
    def _calculate_confidence(self, item: dict, row_position: int, total_results: int) -> int:
        """Рассчитывает процент уверенности бота в результате"""
        try:
            # Базовый балл на основе позиции в результатах
            position_score = max(0, 100 - (row_position - 2) * 8)  # Первый результат = 100%, каждый следующий -8%
            
            # Анализ совпадения названия с запросом
            name = item.get('name', '').lower()
            query = self.user_request.lower()
            
            # Подсчет совпадающих слов
            query_words = [word for word in query.split() if len(word) > 2]
            name_words = name.split()
            
            matching_words = 0
            for q_word in query_words:
                for n_word in name_words:
                    if q_word in n_word or n_word in q_word:
                        matching_words += 1
                        break
            
            # Балл за совпадение слов
            word_match_score = min(30, (matching_words / len(query_words)) * 30) if query_words else 0
            
            # Балл за точность совпадения
            if query in name:
                exact_match_score = 40
            else:
                exact_match_score = 0
            
            # Итоговый балл
            total_confidence = min(100, position_score + word_match_score + exact_match_score)
            
            return int(total_confidence)
            
        except Exception as e:
            # В случае ошибки возвращаем базовый балл
            return max(0, 100 - (row_position - 2) * 8)

    def _generate_clarification_question(self, item: dict, confidence: int) -> str:
        """Генерирует вопрос для уточнения на основе уверенности"""
        try:
            name = item.get('name', '').lower()
            query = self.user_request.lower()
            
            # Простые правила для генерации вопросов
            if confidence < 50:
                return "Это точно то, что вы ищете? Можете уточнить параметры?"
            elif confidence < 70:
                return "Какие характеристики наиболее важны для вас?"
            elif confidence < 90:
                return "Подходит ли этот вариант или нужны другие параметры?"
            else:
                return ""
                
        except Exception as e:
            return "Можете уточнить параметры?"

    def _extract_standard(self, name: str) -> str:
        """Извлекает стандарт из названия (DIN 933, DIN 603 и т.д.)"""
        if not name:
            return ""
        
        import re
        # Ищем DIN стандарты
        din_match = re.search(r'DIN\s*(\d+)', name, re.IGNORECASE)
        if din_match:
            return f"DIN {din_match.group(1)}"
        
        # Ищем другие стандарты
        other_standards = re.findall(r'(ГОСТ|ISO|EN)\s*(\d+)', name, re.IGNORECASE)
        if other_standards:
            return f"{other_standards[0][0]} {other_standards[0][1]}"
        
        return ""

    def _extract_strength_class(self, name: str) -> str:
        """Извлекает класс прочности из названия (кл.пр.8.8, A2-70 и т.д.)"""
        if not name:
            return ""
        
        import re
        # Ищем класс прочности (кл.пр.8.8, кл.пр.10.9)
        strength_match = re.search(r'кл\.пр\.(\d+\.\d+)', name, re.IGNORECASE)
        if strength_match:
            return f"кл.пр.{strength_match.group(1)}"
        
        # Ищем другие обозначения прочности (A2-70, A4-80)
        other_strength = re.search(r'([A-Z]\d+-\d+)', name)
        if other_strength:
            return other_strength.group(1)
        
        return ""

    async def generate_search_results_excel(self, search_results: list, user_intent: dict, original_query: str) -> str:
        """Генерирует Excel файл с результатами поиска"""
        try:
            # Создаем временный файл
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            
            # Создаем рабочую книгу
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Результаты поиска"
            
            # Заголовки
            headers = [
                "№", "Артикул", "Наименование", "Тип", "Диаметр", "Длина", 
                "Материал", "Покрытие", "Стандарт", "Класс прочности", 
                "Количество в упаковке", "Цена", "Примечания"
            ]
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Заполняем данными
            for row, result in enumerate(search_results, 2):
                worksheet.cell(row=row, column=1, value=row-1)  # №
                worksheet.cell(row=row, column=2, value=result.get('sku', ''))  # Артикул
                worksheet.cell(row=row, column=3, value=result.get('name', ''))  # Наименование
                worksheet.cell(row=row, column=4, value=result.get('type', ''))  # Тип
                worksheet.cell(row=row, column=5, value=result.get('diameter', ''))  # Диаметр
                worksheet.cell(row=row, column=6, value=result.get('length', ''))  # Длина
                worksheet.cell(row=row, column=7, value=result.get('material', ''))  # Материал
                worksheet.cell(row=row, column=8, value=result.get('coating', ''))  # Покрытие
                worksheet.cell(row=row, column=9, value=result.get('standard', ''))  # Стандарт
                worksheet.cell(row=row, column=10, value=result.get('strength_class', ''))  # Класс прочности
                worksheet.cell(row=row, column=11, value=result.get('pack_quantity', ''))  # Количество в упаковке
                worksheet.cell(row=row, column=12, value=result.get('price', ''))  # Цена
                worksheet.cell(row=row, column=13, value=result.get('notes', ''))  # Примечания
            
            # Автоматически подгоняем ширину колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            workbook.save(temp_file.name)
            
            logger.info(f"Excel файл создан: {temp_file.name}")
            return temp_file.name
            
        except Exception as e:
            logger.error(f"Ошибка при создании Excel файла: {e}")
            raise

