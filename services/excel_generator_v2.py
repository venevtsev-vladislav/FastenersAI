"""
Enhanced Excel generator for the new FastenersAI architecture
Based on the comprehensive specification
"""

import logging
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pipeline.processing_pipeline import ProcessingResult

logger = logging.getLogger(__name__)

class ExcelGeneratorV2:
    """Enhanced Excel generator with multiple sheets according to specification"""
    
    def __init__(self):
        self.workbook = None
        self.sheets = {}
    
    async def generate_excel(self, request_id: str, results: List[ProcessingResult], 
                           request_data: Dict) -> str:
        """Generate Excel file with multiple sheets"""
        try:
            # Create workbook
            self.workbook = Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create sheets
            self._create_summary_sheet(results)
            self._create_candidates_sheet(results)
            self._create_input_sheet(request_data)
            self._create_errors_sheet(results)
            
            # Apply formatting
            self._apply_formatting()
            
            # Create temp file
            temp_file = self._create_temp_file(request_id)
            
            # Save workbook
            self.workbook.save(temp_file)
            logger.info(f"Excel file generated: {temp_file}")
            
            return temp_file
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            raise
    
    def _create_summary_sheet(self, results: List[ProcessingResult]):
        """Create 'Итог' sheet with final results"""
        sheet = self.workbook.create_sheet("Итог")
        self.sheets['summary'] = sheet
        
        # Headers
        headers = [
            "№", "Запрос", "KU", "Наименование", "Упаковка_pack_qty",
            "Кол-во_уп", "Кол-во_шт", "Цена", "Сумма", "Статус"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row, result in enumerate(results, 2):
            # Get item details if KU is chosen
            item_name = ""
            pack_qty = ""
            price = ""
            
            if result.chosen_ku:
                # This would normally come from database lookup
                item_name = f"Item for {result.chosen_ku}"
                pack_qty = "100"  # Placeholder
                price = "10.50"   # Placeholder
            
            sheet.cell(row=row, column=1, value=row - 1)  # №
            sheet.cell(row=row, column=2, value=result.raw_text)  # Запрос
            sheet.cell(row=row, column=3, value=result.chosen_ku or "")  # KU
            sheet.cell(row=row, column=4, value=item_name)  # Наименование
            sheet.cell(row=row, column=5, value=pack_qty)  # Упаковка_pack_qty
            sheet.cell(row=row, column=6, value=result.qty_packs or "")  # Кол-во_уп
            sheet.cell(row=row, column=7, value=result.qty_units or "")  # Кол-во_шт
            sheet.cell(row=row, column=8, value=price)  # Цена
            sheet.cell(row=row, column=9, value=result.total or "")  # Сумма
            sheet.cell(row=row, column=10, value=result.status)  # Статус
    
    def _create_candidates_sheet(self, results: List[ProcessingResult]):
        """Create 'Кандидаты' sheet with all candidates"""
        sheet = self.workbook.create_sheet("Кандидаты")
        self.sheets['candidates'] = sheet
        
        # Headers
        headers = [
            "№ позиции", "Сырой запрос", "Кандидат KU", "Имя", "score",
            "pack_qty", "price", "explanation", "source"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row = 2
        for result in results:
            for candidate in result.candidates:
                sheet.cell(row=row, column=1, value=result.line_id)  # № позиции
                sheet.cell(row=row, column=2, value=result.raw_text)  # Сырой запрос
                sheet.cell(row=row, column=3, value=candidate.ku)  # Кандидат KU
                sheet.cell(row=row, column=4, value=candidate.name)  # Имя
                sheet.cell(row=row, column=5, value=candidate.score)  # score
                sheet.cell(row=row, column=6, value=candidate.pack_qty or "")  # pack_qty
                sheet.cell(row=row, column=7, value=candidate.price or "")  # price
                sheet.cell(row=row, column=8, value=candidate.explanation)  # explanation
                sheet.cell(row=row, column=9, value=candidate.source)  # source
                row += 1
    
    def _create_input_sheet(self, request_data: Dict):
        """Create 'Вход' sheet with input data"""
        sheet = self.workbook.create_sheet("Вход")
        self.sheets['input'] = sheet
        
        # Headers
        headers = ["№ позиции", "Сырой запрос", "source", "attachment_id"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        input_lines = request_data.get('input_lines', [])
        for row, line in enumerate(input_lines, 2):
            sheet.cell(row=row, column=1, value=row - 1)  # № позиции
            sheet.cell(row=row, column=2, value=line)  # Сырой запрос
            sheet.cell(row=row, column=3, value=request_data.get('source', 'text'))  # source
            sheet.cell(row=row, column=4, value=request_data.get('attachment_id', ''))  # attachment_id
    
    def _create_errors_sheet(self, results: List[ProcessingResult]):
        """Create 'Ошибки' sheet with error information"""
        sheet = self.workbook.create_sheet("Ошибки")
        self.sheets['errors'] = sheet
        
        # Headers
        headers = ["№ позиции", "Сырой запрос", "причина", "подсказка"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data - only error results
        row = 2
        for result in results:
            if result.status in ['error', 'not_found']:
                sheet.cell(row=row, column=1, value=result.line_id)  # № позиции
                sheet.cell(row=row, column=2, value=result.raw_text)  # Сырой запрос
                
                if result.status == 'error':
                    sheet.cell(row=row, column=3, value="Ошибка обработки")  # причина
                    sheet.cell(row=row, column=4, value="Проверьте формат запроса")  # подсказка
                elif result.status == 'not_found':
                    sheet.cell(row=row, column=3, value="Не найдено совпадений")  # причина
                    sheet.cell(row=row, column=4, value="Уточните параметры или добавьте в каталог")  # подсказка
                
                row += 1
    
    def _apply_formatting(self):
        """Apply formatting to all sheets"""
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to all sheets
        for sheet_name, sheet in self.sheets.items():
            # Apply borders to all cells
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format currency columns
            if sheet_name == 'summary':
                # Format price and total columns
                for row in range(2, sheet.max_row + 1):
                    # Price column (H)
                    price_cell = sheet.cell(row=row, column=8)
                    if price_cell.value:
                        price_cell.number_format = '#,##0.00'
                    
                    # Total column (I)
                    total_cell = sheet.cell(row=row, column=9)
                    if total_cell.value:
                        total_cell.number_format = '#,##0.00'
            
            # Format score column in candidates sheet
            if sheet_name == 'candidates':
                for row in range(2, sheet.max_row + 1):
                    score_cell = sheet.cell(row=row, column=5)
                    if score_cell.value:
                        score_cell.number_format = '0.000'
    
    def _create_temp_file(self, request_id: str) -> str:
        """Create temporary file for Excel"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.gettempdir()
        filename = f"fasteners_result_{request_id}_{timestamp}.xlsx"
        return os.path.join(temp_dir, filename)

# Global Excel generator instance
_excel_generator_v2 = None

def get_excel_generator_v2() -> ExcelGeneratorV2:
    """Get global Excel generator instance"""
    global _excel_generator_v2
    if _excel_generator_v2 is None:
        _excel_generator_v2 = ExcelGeneratorV2()
    return _excel_generator_v2
