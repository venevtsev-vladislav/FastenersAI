#!/usr/bin/env python3
"""
Тест генерации Excel файла для отладки проблемы
"""

import asyncio
import sys
import os

# Добавляем родительскую папку в путь для импорта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from services.excel_generator import ExcelGenerator

async def test_excel_generation():
    """Тестируем генерацию Excel файла с тестовыми данными"""
    
    print("🔍 ТЕСТ ГЕНЕРАЦИИ EXCEL ФАЙЛА")
    print("=" * 50)
    
    # Тестовые данные - имитируем результат поиска
    test_search_results = [
        {
            "order_position": 1,
            "search_query": "саморез по металлу 4,2х25",
            "diameter": "4.2",
            "length": "25",
            "material": "сталь",
            "coating": "оцинкованный",
            "requested_quantity": 100,
            "confidence": 0.95,
            "sku": "TEST-001",
            "name": "Саморез по металлу 4,2х25 оцинкованный",
            "smart_probability": 0.92,
            "pack_size": 1000,
            "unit": "шт"
        },
        {
            "order_position": 2,
            "search_query": "саморез по дереву 4,2х25",
            "diameter": "4.2",
            "length": "25",
            "material": "сталь",
            "coating": "фосфатированный",
            "requested_quantity": 50,
            "confidence": 0.88,
            "sku": "TEST-002",
            "name": "Саморез по дереву 4,2х25 фосфатированный",
            "smart_probability": 0.85,
            "pack_size": 500,
            "unit": "шт"
        }
    ]
    
    user_request = "саморезы 4,2х25"
    
    try:
        # Создаем генератор Excel
        excel_generator = ExcelGenerator()
        
        print("✅ Excel генератор создан")
        print(f"📊 Тестовые данные: {len(test_search_results)} позиций")
        
        # Генерируем Excel файл
        excel_file = await excel_generator.generate_excel(
            search_results=test_search_results,
            user_request=user_request
        )
        
        print(f"✅ Excel файл создан: {excel_file}")
        print(f"📁 Размер файла: {os.path.getsize(excel_file)} байт")
        
        # Проверяем содержимое файла
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📋 Содержимое Excel файла:")
        print(f"  - Лист: {ws.title}")
        print(f"  - Строк: {ws.max_row}")
        print(f"  - Колонок: {ws.max_column}")
        
        # Показываем заголовки
        print(f"\n📝 Заголовки:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"  {col}: {header}")
        
        # Показываем данные
        print(f"\n📊 Данные:")
        for row in range(2, ws.max_row + 1):
            print(f"  Строка {row}:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                print(f"    {col}: {value}")
        
        return True
        
    except Exception as e:
        print(f"💥 ОШИБКА при генерации Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Запуск теста генерации Excel...")
    
    # Запускаем асинхронный тест
    result = asyncio.run(test_excel_generation())
    
    if result:
        print("\n🎉 ТЕСТ ГЕНЕРАЦИИ EXCEL ПРОЙДЕН!")
    else:
        print("\n❌ ТЕСТ ГЕНЕРАЦИИ EXCEL НЕ ПРОЙДЕН!")
    
    sys.exit(0 if result else 1)
