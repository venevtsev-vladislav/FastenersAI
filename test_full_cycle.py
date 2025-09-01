#!/usr/bin/env python3
"""
Тест полного цикла работы бота с исправленной вероятностью
"""

import asyncio
import logging
from database.supabase_client import search_parts, init_supabase
from services.excel_generator import ExcelGenerator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

async def test_full_cycle():
    """Тестирует полный цикл: поиск -> Excel генерация"""
    
    # Инициализируем Supabase
    print("🔧 Инициализирую Supabase...")
    await init_supabase()
    
    # Тестовые данные из логов пользователя
    test_query = "винт мебельный 640 оцинкованный"
    test_user_intent = {
        "type": "винт мебельный",
        "diameter": "6",
        "length": "40",
        "material": None,
        "coating": "оцинкованный",
        "quantity": "200 шт",
        "confidence": 0.9
    }
    
    print("🔍 Тестирую поиск через Supabase...")
    
    # Поиск деталей
    search_results = await search_parts(test_query, test_user_intent)
    
    print(f"✅ Найдено {len(search_results)} результатов")
    
    # Анализируем результаты
    for i, result in enumerate(search_results[:5]):
        print(f"\n--- Результат {i+1} ---")
        print(f"Название: {result.get('name', 'N/A')}")
        print(f"SKU: {result.get('sku', 'N/A')}")
        print(f"Вероятность: {result.get('probability_percent', 'N/A')}%")
        print(f"Релевантность: {result.get('relevance_score', 'N/A')}")
        print(f"Причина совпадения: {result.get('match_reason', 'N/A')}")
    
    # Тестируем генерацию Excel
    print(f"\n📊 Тестирую генерацию Excel...")
    
    excel_generator = ExcelGenerator()
    excel_file = await excel_generator.generate_excel(search_results, test_query)
    
    print(f"✅ Excel файл создан: {excel_file}")
    
    # Проверяем содержимое Excel (читаем первые несколько строк)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📋 Содержимое Excel файла:")
        print(f"Заголовки: {[ws.cell(row=1, column=i).value for i in range(1, 13)]}")
        
        # Показываем первые 3 строки данных
        for row in range(2, min(5, ws.max_row + 1)):
            print(f"\nСтрока {row-1}:")
            print(f"  Наименование: {ws.cell(row=row, column=7).value}")
            print(f"  Релевантность: {ws.cell(row=row, column=11).value}")
            print(f"  Вероятность: {ws.cell(row=row, column=12).value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка при чтении Excel: {e}")

if __name__ == "__main__":
    asyncio.run(test_full_cycle())
